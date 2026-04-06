#!/usr/bin/env python3
"""
SkillsFuture Dashboard  ─  Flask Backend
=========================================
Routes:
  GET  /                       → Main UI
  POST /api/scrape/start        → Launch scraping job
  GET  /api/scrape/progress/<id>→ SSE stream for real-time progress
  GET  /api/history             → List all past searches
  GET  /api/data/<keyword>      → Return JSON data for a keyword
  POST /api/data/delete         → Delete a keyword's data
  GET  /api/export/<keyword>    → Download Excel for keyword
"""

import sys, asyncio, re, json, uuid, threading, time, io, os
from pathlib import Path
from datetime import datetime, timezone
from collections import defaultdict

# ── Windows asyncio fix ───────────────────────────────────────
if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

try:
    import nest_asyncio
    nest_asyncio.apply()
except ImportError:
    pass

from flask import (Flask, render_template, request, jsonify,
                   Response, send_file, abort)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── App setup ─────────────────────────────────────────────────
app = Flask(__name__)
DATA_DIR = Path("skillsfuture_data")
DATA_DIR.mkdir(exist_ok=True)
HISTORY_FILE = DATA_DIR / "history.json"

# In-memory job tracker  { job_id: { status, pages_done, total_pages, ... } }
JOBS: dict[str, dict] = {}
JOBS_LOCK = threading.Lock()

# ─────────────────────────────────────────────────────────────
# HISTORY HELPERS
# ─────────────────────────────────────────────────────────────

def keyword_to_filename(keyword: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", keyword).replace(" ", "_").lower()

def excel_path(keyword: str) -> Path:
    return DATA_DIR / f"{keyword_to_filename(keyword)}.xlsx"

# ─────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────

COLUMNS = [
    ("Course Title",               52),
    ("Training Provider",          32),
    ("Course URL",                 40),
    ("Full Course Fee",            18),
    ("After SkillsFuture Funding", 22),
    ("Star Rating",                14),
    ("No. of Ratings",             14),
    ("Upcoming Course Date",       22),
    ("Scraped At",                 26),
]

_thin       = Side(style="thin", color="C0C0C0")
HEADER_FILL = PatternFill("solid", fgColor="1A4E8C")
ALT_FILL    = PatternFill("solid", fgColor="EBF0FA")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="Calibri", size=10)
TITLE_FONT  = Font(name="Calibri", bold=True, size=13, color="1A4E8C")
BORDER      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="top")


def build_workbook(courses: list[dict], keyword: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "SkillsFuture Courses"
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    ws["A1"] = (
        f'SkillsFuture: "{keyword}"  |  '
        f"Exported {datetime.now().strftime('%d %b %Y %H:%M')}"
    )
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22
    for row_idx, course in enumerate(courses, 3):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, (key, _) in enumerate(COLUMNS, 1):
            val = course.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT; cell.alignment = WRAP; cell.border = BORDER
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 40
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}2"
    return wb


def save_courses_to_excel(courses: list[dict], keyword: str, mode: str = "overwrite"):
    """mode: 'overwrite' | 'append'"""
    path = excel_path(keyword)
    existing: list[dict] = []

    if mode == "append" and path.exists():
        # Read existing rows back
        wb_old = load_workbook(path)
        ws_old = wb_old.active
        headers = [ws_old.cell(2, c).value for c in range(1, len(COLUMNS) + 1)]
        for row in ws_old.iter_rows(min_row=3, values_only=True):
            if row[0]:  # Course Title present
                existing.append(dict(zip(headers, row)))

    # Deduplicate by (Course Title + Training Provider)
    seen = {(c.get("Course Title",""), c.get("Training Provider","")) for c in existing}
    new_unique = [
        c for c in courses
        if (c.get("Course Title",""), c.get("Training Provider","")) not in seen
    ]
    all_courses = existing + new_unique

    wb = build_workbook(all_courses, keyword)
    wb.save(path)
    return all_courses, len(new_unique)


def load_courses_from_excel(keyword: str) -> list[dict]:
    path = excel_path(keyword)
    if not path.exists():
        return []
    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(2, c).value for c in range(1, len(COLUMNS) + 1)]
    courses = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:
            courses.append(dict(zip(headers, row)))
    return courses

# ─────────────────────────────────────────────────────────────
# SCRAPER  (integrated from the provided script)
# ─────────────────────────────────────────────────────────────

SEARCH_BASE = (
    "https://www.myskillsfuture.gov.sg/content/portal/en/"
    "portal-search/portal-search.html"
    "?fq=Course_Supp_Period_To_1%3A%5B2026-03-18T00%3A00%3A00Z%20TO%20*%5D"
    "&fq=IsValid%3Atrue&q="
)


async def _scrape_page(page, import_playwright_page) -> list[dict]:
    """Extract all course cards from the currently visible page."""
    await import_playwright_page.wait_for_selector(
        "div.card", timeout=30_000, state="attached"
    )
    await asyncio.sleep(1.5)
    courses = await import_playwright_page.evaluate("""
    () => {
        const results = [];
        for (const card of document.querySelectorAll("div.card")) {
            const providerEl = card.querySelector("div.course-provider a");
            const provider   = providerEl ? providerEl.innerText.trim() : "N/A";
            const titleEl    = card.querySelector(
                "h5.card-title a[data-bind*='courseTitleShorten']");
            const title      = titleEl ? titleEl.innerText.trim() : "";
            if (!title) continue;
            const courseUrl  = titleEl ? (titleEl.href || "") : "";
            const dateEl     = card.querySelector(
                "li.list-group-item strong[data-bind*='formatDate']");
            const upcomingDate = dateEl ? dateEl.innerText.trim() : "None listed";
            const starsHolder  = card.querySelector("div.stars-holder");
            let starRating = null;
            if (starsHolder) {
                const filled = starsHolder.querySelectorAll("i.fa-solid.fa-star").length;
                const half   = starsHolder.querySelectorAll(
                    "i.fa-solid.fa-star-half-stroke, i.fa-solid.fa-star-half").length;
                if (filled > 0 || half > 0) starRating = parseFloat((filled + half * 0.5).toFixed(1));
            }
            const ratingsEl = card.querySelector("span[data-bind*='NumberOfRespondents']");
            let numRatings = "0";
            if (ratingsEl) { const m = ratingsEl.innerText.match(/(\\d+)/); if (m) numRatings = m[1]; }
            const feeEl  = card.querySelector("strong[data-bind*='Tol_Cost_of_Trn_Per_Trainee']");
            const fullFee = feeEl ? feeEl.innerText.trim() : "N/A";
            const nettDiv = card.querySelector("div[name='nettfee']");
            let nettFee = "N/A";
            if (nettDiv) { const nettEl = nettDiv.querySelector("strong"); if (nettEl) nettFee = nettEl.innerText.trim(); }
            results.push({ title, provider, courseUrl, fullFee, nettFee,
                           starRating, numRatings, upcomingDate });
        }
        return results;
    }
    """)
    return [
        {
            "Course Title":                 c["title"],
            "Training Provider":            c["provider"],
            "Course URL":                   c["courseUrl"],
            "Full Course Fee":              c["fullFee"],
            "After SkillsFuture Funding":   c["nettFee"],
            "Star Rating":                  c["starRating"],
            "No. of Ratings":               int(c["numRatings"]),
            "Upcoming Course Date":         c["upcomingDate"],
            "Scraped At":                   datetime.now(timezone.utc).isoformat(),
        }
        for c in courses if c.get("title")
    ]


async def _run_scraper(job_id: str, keyword: str, max_pages: int, mode: str):
    """
    Core async scraper. Updates JOBS[job_id] as it progresses.
    Saves results to Excel and updates history on completion.
    """
    import urllib.parse
    from playwright.async_api import async_playwright

    url = SEARCH_BASE + urllib.parse.quote(keyword)
    all_courses: list[dict] = []
    seen_titles: set[str]   = set()
    warnings: list[str]     = []

    def _set(key, val):
        with JOBS_LOCK:
            JOBS[job_id][key] = val

    _set("status",      "running")
    _set("pages_done",  0)
    _set("total_pages", max_pages)
    _set("courses",     [])
    _set("warnings",    [])
    _set("error",       None)

    # Detect scrape-all mode (sentinel 999999)
    scrape_all = max_pages >= 999999

    try:
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True)
            context = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1280, "height": 900},
            )
            page = await context.new_page()

            _set("message", f"Opening MySkillsFuture for '{keyword}'…")
            await page.goto(url, wait_until="networkidle", timeout=90_000)
            await asyncio.sleep(3)

            for page_num in range(1, max_pages + 1):
                _set("pages_done", page_num - 1)
                if scrape_all:
                    _set("message", f"Scraping page {page_num} (scraping all pages)…")
                    _set("total_pages", page_num)   # keep total in sync so UI shows real count
                else:
                    _set("message", f"Scraping page {page_num} of {max_pages}…")

                try:
                    page_courses = await _scrape_page(page, page)
                except Exception as e:
                    warnings.append(f"Page {page_num}: {str(e)[:120]}")
                    _set("warnings", warnings)
                    break

                new = [c for c in page_courses
                       if c["Course Title"] not in seen_titles]
                seen_titles.update(c["Course Title"] for c in new)
                all_courses.extend(new)
                _set("courses", all_courses)

                # Check for next page
                next_btn = await page.query_selector(
                    'a.page-link[aria-label="View next page"]'
                )
                if not next_btn:
                    break  # no pagination — single page
                parent_li    = await next_btn.evaluate_handle("el => el.closest('li')")
                parent_class = await (await parent_li.get_property("className")).json_value()
                if "disabled" in parent_class:
                    break

                await next_btn.click()
                await page.wait_for_load_state("networkidle")
                await asyncio.sleep(2)

            _set("pages_done", page_num)
            await browser.close()

        if not all_courses:
            _set("status",  "no_results")
            _set("message", f"No courses found for '{keyword}'.")
            return

        # Save to Excel
        all_saved, added = save_courses_to_excel(all_courses, keyword, mode)


        _set("status",  "completed")
        _set("message", f"Done! {added} new courses added ({len(all_saved)} total).")
        _set("courses", all_saved)
        _set("warnings", warnings)

    except Exception as e:
        _set("status",  "error")
        _set("error",   str(e))
        _set("message", f"Scraping failed: {str(e)[:200]}")


def _thread_scrape(job_id, keyword, max_pages, mode):
    """Run the async scraper in a dedicated thread with its own event loop."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(_run_scraper(job_id, keyword, max_pages, mode))
    finally:
        loop.close()

# ─────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/data/<keyword>")
def api_data(keyword):
    courses = load_courses_from_excel(keyword)
    if not courses:
        return jsonify({"error": "No data found"}), 404
    return jsonify(courses)


@app.route("/api/check/<keyword>")
def api_check(keyword):
    """Return whether keyword already has stored data."""
    path = excel_path(keyword)
    if path.exists():
        courses = load_courses_from_excel(keyword)
        return jsonify({"exists": True, "count": len(courses)})
    return jsonify({"exists": False})


@app.route("/api/scrape/start", methods=["POST"])
def api_scrape_start():
    body      = request.get_json(force=True)
    keyword   = (body.get("keyword") or "").strip()
    raw_pages = body.get("max_pages", 5)
    mode      = body.get("mode", "overwrite")  # overwrite | append

    # 0 or empty → scrape ALL pages (sentinel = 999999)
    try:
        max_pages = int(raw_pages)
    except (TypeError, ValueError):
        max_pages = 0

    scrape_all = max_pages <= 0
    if scrape_all:
        max_pages = 999999   # effectively unlimited; loop breaks on last page

    if not keyword:
        return jsonify({"error": "Keyword is required."}), 400
    if not scrape_all and max_pages > 200:
        return jsonify({"error": "MAX_PAGES must be between 1 and 200 (or 0 for all)."}), 400
    if mode not in ("overwrite", "append"):
        return jsonify({"error": "Invalid mode."}), 400

    job_id = str(uuid.uuid4())
    with JOBS_LOCK:
        JOBS[job_id] = {
            "status": "queued", "keyword": keyword,
            "pages_done": 0, "total_pages": max_pages,
            "scrape_all": scrape_all,
            "message": "Starting…", "courses": [],
            "warnings": [], "error": None,
        }

    t = threading.Thread(
        target=_thread_scrape, args=(job_id, keyword, max_pages, mode), daemon=True
    )
    t.start()
    return jsonify({"job_id": job_id})


@app.route("/api/scrape/progress/<job_id>")
def api_scrape_progress(job_id):
    """Server-Sent Events stream — pushes job state every second."""
    def generate():
        last_ping = time.time()
        while True:
            with JOBS_LOCK:
                job = JOBS.get(job_id)
            if not job:
                yield f"data: {json.dumps({'error': 'Job not found'})}\n\n"
                break
            payload = {
                "status":     job["status"],
                "pages_done": job["pages_done"],
                "total":      job["total_pages"],
                "scrape_all": job.get("scrape_all", False),
                "message":    job["message"],
                "count":      len(job.get("courses", [])),
                "warnings":   job.get("warnings", []),
                "error":      job.get("error"),
            }
            yield f"data: {json.dumps(payload)}\n\n"
            if job["status"] in ("completed", "error", "no_results"):
                break
            # Send a comment ping every 15s to prevent proxy timeout
            if time.time() - last_ping > 15:
                yield ": ping\n\n"
                last_ping = time.time()
            time.sleep(1)

    resp = Response(generate(), mimetype="text/event-stream")
    resp.headers["Cache-Control"] = "no-cache, no-transform"
    resp.headers["X-Accel-Buffering"] = "no"
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["Connection"] = "keep-alive"
    return resp


@app.route("/api/export/<keyword>")
def api_export(keyword):
    path = excel_path(keyword)
    if not path.exists():
        abort(404)
    return send_file(
        path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"skillsfuture_{keyword_to_filename(keyword)}.xlsx",
    )


@app.route("/api/scrape/progress-check/<job_id>")
def api_scrape_progress_check(job_id):
    """Single-shot status check — used by frontend reconnect logic."""
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify({"status": "error", "error": "Job not found"})
    return jsonify({
        "status":     job["status"],
        "pages_done": job["pages_done"],
        "total":      job["total_pages"],
        "scrape_all": job.get("scrape_all", False),
        "message":    job["message"],
        "count":      len(job.get("courses", [])),
        "warnings":   job.get("warnings", []),
        "error":      job.get("error"),
    })



if __name__ == "__main__":
    print("SkillsFuture Dashboard running at http://127.0.0.1:5000")
    app.run(debug=True, threaded=True)
